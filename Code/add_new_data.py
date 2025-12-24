
import os
from openpyxl import load_workbook

# Get the absolute path of the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))

def add_data_to_excel(file_path_txt, label):
    # --- Parse Sentences ---
    # Read the file
    with open(file_path_txt, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Extract sentences
    sentences = []
    for line in lines:
        stripped_line = line.strip()
        if stripped_line:
            sentences.append(stripped_line)

    # --- Add to Excel ---
    # Dataset yolu
    file_path_xlsx = os.path.join(os.path.dirname(script_dir), "hatespeech_dataset.xlsx")

    # Excel dosyasını yükle
    book = load_workbook(file_path_xlsx)

    # "Dengeli Veriseti" sayfasına yeni verileri ekle
    sheet = book["Dengeli Veriseti"]

    # Get the last row number
    last_row_num = sheet.max_row
    last_row_id_num = int(sheet.cell(row=last_row_num, column=1).value.replace("Row", ""))

    # Yeni verileri oluştur
    new_data = []
    for i, sentence in enumerate(sentences):
        new_row_id = f"Row{last_row_id_num + i + 1}"
        new_data.append((new_row_id, sentence, label, ""))

    # Verileri ekle
    for row in new_data:
        sheet.append(row)

    # Değişiklikleri kaydet
    book.save(file_path_xlsx)

    print(f"{len(sentences)} yeni veri ('{label}') 'Dengeli Veriseti' sayfasına eklendi.")

# Add positive sentences
positive_sentences_path = os.path.join(script_dir, "positive_sentences.txt")
add_data_to_excel(positive_sentences_path, "hiçbiri")

# Add saldırgan sentences
saldırgan_sentences_path = os.path.join(script_dir, "saldırgan_sentences.txt")
add_data_to_excel(saldırgan_sentences_path, "saldırgan")

# Add tehdit sentences
tehdit_sentences_path = os.path.join(script_dir, "tehdit_sentences.txt")
add_data_to_excel(tehdit_sentences_path, "tehdit")
