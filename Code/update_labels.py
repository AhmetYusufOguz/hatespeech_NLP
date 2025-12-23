
from openpyxl import load_workbook
import os

# Get the absolute path of the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)

# Dataset yolu
file_path = os.path.join(project_root, "hatespeech_dataset.xlsx")

# Excel dosyasını yükle
book = load_workbook(file_path)

# "Dengeli Veriseti" sayfasını seç
sheet = book["Dengeli Veriseti"]

# "saldırgan" etiketini "nefret" olarak değiştir
for row in sheet.iter_rows(min_row=2): # Başlık satırını atla
    if row[2].value == "saldırgan":
        row[2].value = "nefret"

# Değişiklikleri kaydet
book.save(file_path)

print("'saldırgan' etiketleri 'nefret' olarak değiştirildi.")
