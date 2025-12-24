
import os
from openpyxl import load_workbook

# Get the absolute path of the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))

# --- Parse Sentences ---
# Read the file
file_path_txt = os.path.join(script_dir, "Niyet cümleleri.txt")
with open(file_path_txt, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# Find the start of the sentences
start_index = -1
for i, line in enumerate(lines):
    if "Eklenen 50 Yeni Cümle:" in line:
        start_index = i + 1
        break

# Extract sentences
sentences = []
if start_index != -1:
    for line in lines[start_index:]:
        stripped_line = line.strip()
        if stripped_line:
            sentences.append(stripped_line)

# --- Add to Excel ---
# Dataset yolu
file_path_xlsx = os.path.join(os.path.dirname(script_dir), "hatespeech_dataset.xlsx")

# Excel dosyasını yükle
book = load_workbook(file_path_xlsx)

# "Dengeli Veriseti" sayfasına yeni verileri ekle
sheet = book["Dengeli Veriseti"]

# Get the last row number
last_row_num = sheet.max_row
last_row_id_num = int(sheet.cell(row=last_row_num, column=1).value.replace("Row", ""))

# Yeni verileri oluştur
new_data = []
for i, sentence in enumerate(sentences):
    new_row_id = f"Row{last_row_id_num + i + 1}"
    new_data.append((new_row_id, sentence, "niyet", ""))

# Verileri ekle
for row in new_data:
    sheet.append(row)

# Değişiklikleri kaydet
book.save(file_path_xlsx)

print(f"{len(sentences)} yeni veri 'Dengeli Veriseti' sayfasına eklendi.")
